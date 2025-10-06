import streamlit as st
import pandas as pd
import requests
from bs4 import BeautifulSoup
import re
import logging
from urllib.parse import urljoin, urlparse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
import os
import json
from dotenv import load_dotenv
import time


import streamlit as st
import google.generativeai as genai
load_dotenv()

api_key = None
source = None

# Try Streamlit secrets only if available
try:
    api_key = st.secrets.get("GOOGLE_API_KEY")
    if api_key:
        source = "Streamlit Secrets"
except Exception:
    # st.secrets not available locally
    pass

# Fallback to .env if secrets not found
if not api_key:
    api_key = os.getenv("GOOGLE_API_KEY")
    if api_key:
        source = ".env file"

# Strip whitespace if key exists
if api_key:
    api_key = api_key.strip()

# Configure Gemini or show error if missing
if api_key:
    genai.configure(api_key=api_key)
    st.success(f"✅ Gemini API key loaded successfully from {source}")
else:
    st.error(
        "⚠️ Gemini API key not found! Please set it in Streamlit Secrets (cloud) or .env file (local)."
    )
#logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

if 'scraped_data' not in st.session_state:    #session state
    st.session_state.scraped_data = []
if 'current_url' not in st.session_state:
    st.session_state.current_url = ""

st.markdown("""
<style>
    .main-header {
        font-size: 2.5rem;
        font-weight: 700;
        color: #1e88e5;
        text-align: center;
        margin-bottom: 2rem;
    }
    .stButton>button {
        background-color: #1e88e5;
        color: white;
        border-radius: 8px;
        padding: 0.5rem 1rem;
        font-weight: 600;
    }
    .stButton>button:hover {
        background-color: #1976d2;
    }
    .data-card {
        background-color: #f5f9ff;
        border-radius: 8px;
        padding: 1.2rem;
        margin: 0.8rem 0;
        border-left: 4px solid #4caf50;
    }
    .field-label {
        font-weight: 600;
        color: #1e88e5;
        margin-bottom: 0.3rem;
    }
    .field-value {
        color: #333;
        line-height: 1.5;
        font-size: 1.1em;
    }
    .not-found {
        color: #f44336 !important;
    }
    .ai-badge {
        background-color: #e8f5e9;
        color: #2e7d32;
        padding: 0.2rem 0.6rem;
        border-radius: 12px;
        font-size: 0.85rem;
        font-weight: 600;
        margin-left: 0.5rem;
    }
    .success-badge {
        background-color: #4caf50;
        color: white;
        padding: 0.2rem 0.6rem;
        border-radius: 12px;
        font-size: 0.85rem;
        font-weight: 600;
    }
    .retry-info {
        background-color: #fff3cd;
        border-left: 4px solid #ffc107;
        padding: 1rem;
        margin: 1rem 0;
        border-radius: 4px;
    }
</style>
""", unsafe_allow_html=True)

def create_excel_file(data):
    """Create a formatted Excel file from the scraped data"""
    if not data:
        return None
    
    df = pd.DataFrame(data)
    
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='NGO Data')
        
        workbook = writer.book
        worksheet = writer.sheets['NGO Data']
        
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1e88e5", end_color="1e88e5", fill_type="solid")
        alignment = Alignment(horizontal="center", vertical="center")
        
        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment
        
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    output.seek(0)
    return output

def find_contact_pages(base_url, soup):
    """Find URLs of contact, about, and other relevant pages"""
    contact_pages = []
    base_domain = urlparse(base_url).netloc
    
    keywords = ['contact', 'about', 'reach', 'connect', 'get-in-touch', 'office', 'location', 'team']
    
    for link in soup.find_all('a', href=True):
        href = link['href']
        full_url = urljoin(base_url, href)
        
        if urlparse(full_url).netloc == base_domain:
            link_text = link.get_text().lower()
            href_lower = href.lower()
            
            if any(keyword in link_text or keyword in href_lower for keyword in keywords):
                if full_url not in contact_pages and full_url != base_url:
                    contact_pages.append(full_url)
    
    return contact_pages[:3]  

def scrape_comprehensive_content(url):
    """Scrape content from main page AND contact/about pages"""
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
    }
    
    all_content = []
    
    try:
        if not url.startswith(('http://', 'https://')):           #if missing
            url = 'https://' + url   
        
        response = requests.get(url, headers=headers, timeout=15)   # Scrapes main page

        response.raise_for_status()
        
        soup = BeautifulSoup(response.content, 'html.parser')
        
        additional_pages = find_contact_pages(url, soup)
        
        main_soup = BeautifulSoup(response.content, 'html.parser')
        for script in main_soup(["script", "style", "nav", "header", "aside", "img"]):
            script.decompose()
        
        main_text = main_soup.get_text()
        main_text = re.sub(r'\s+', ' ', main_text).strip()
        all_content.append(("Main Page", main_text[:8000]))
        
        for page_url in additional_pages:
            try:
                time.sleep(0.5)  # Be polite to the server
                page_response = requests.get(page_url, headers=headers, timeout=10)
                page_response.raise_for_status()
                
                page_soup = BeautifulSoup(page_response.content, 'html.parser')
                for script in page_soup(["script", "style", "nav", "header", "aside", "img"]):
                    script.decompose()
                
                page_text = page_soup.get_text()
                page_text = re.sub(r'\s+', ' ', page_text).strip()
                all_content.append((page_url.split('/')[-1], page_text[:8000]))
                
            except Exception as e:
                logger.warning(f"Could not scrape {page_url}: {str(e)}")
        
        structured_data = extract_structured_data(soup)
        if structured_data:
            all_content.append(("Structured Data", structured_data))
        
        return all_content, url
        
    except Exception as e:
        logger.error(f"Error scraping {url}: {str(e)}")
        return [("Error", f"Error scraping website: {str(e)}")], url

def extract_structured_data(soup):
    """Extract contact info from footer, contact sections, and metadata"""
    structured_info = []
    
    #searches footer
    footer = soup.find('footer')
    if footer:
        footer_text = footer.get_text()
        footer_text = re.sub(r'\s+', ' ', footer_text).strip()
        structured_info.append(f"FOOTER: {footer_text}")
    
    #get contact
    contact_sections = soup.find_all(['div', 'section'], class_=re.compile(r'contact|address|info', re.I))
    for section in contact_sections[:3]:
        section_text = section.get_text()
        section_text = re.sub(r'\s+', ' ', section_text).strip()
        structured_info.append(f"CONTACT SECTION: {section_text}")

    for link in soup.find_all(['a']):
        href = link.get('href', '')
        if href.startswith('tel:'):
            structured_info.append(f"PHONE: {href.replace('tel:', '')}")
        elif href.startswith('mailto:'):
            structured_info.append(f"EMAIL: {href.replace('mailto:', '')}")
    
    return ' | '.join(structured_info) if structured_info else None

def clean_json_response(response_text):
    """Aggressively clean and extract JSON from AI response"""
    response_text = response_text.strip()         #remove markdown blocks

    
    # Remove ```json or ``` at start
    if response_text.startswith('```json'):
        response_text = response_text[7:]
    elif response_text.startswith('```'):
        response_text = response_text[3:]
    
    
    if response_text.endswith('```'):        # Remove ``` at end
        response_text = response_text[:-3]
    
    response_text = response_text.strip()
    
    #find jsons object
    json_pattern = r'\{[^{}]*(?:\{[^{}]*\}[^{}]*)*\}'
    matches = re.findall(json_pattern, response_text, re.DOTALL)
    
    if matches:
        for match in matches:
            try:
                parsed = json.loads(match)
                if isinstance(parsed, dict) and len(parsed) > 0:
                    return parsed
            except:
                continue
    
    #direct parsing if regex fail
    try:
        return json.loads(response_text)
    except:
        pass
    
    #try to extract keyvalue pairs
    try:
        result = {}
        lines = response_text.split('\n')
        for line in lines:
            if ':' in line and ('"' in line or "'" in line):
                parts = line.split(':', 1)
                if len(parts) == 2:
                    key = parts[0].strip().strip('"').strip("'").strip(',')
                    value = parts[1].strip().strip('"').strip("'").strip(',')
                    if key and value:
                        result[key] = value
        
        if len(result) > 0:
            return result
    except:
        pass
    
    return None

def extract_required_fields_with_gemini(all_content, url, retry_count=0):
    """Use Gemini AI with enhanced multi-page content and retry logic"""
    GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")
    
    if not GEMINI_API_KEY:
        return {
            "NGO Name": "  Gemini API key not configured",
            "Address": "Add GEMINI_API_KEY to .env file",
            "Services Offered": "Not available",
            "Contact Person Details": "Not available",
            "Contact Number": "Not available"
        }
    
    try:
        import google.generativeai as genai
        genai.configure(api_key=GEMINI_API_KEY)
        
        # combines all content
        combined_content = "\n\n=== WEBSITE SECTIONS ===\n\n"
        for source, content in all_content:
            combined_content += f"\n--- {source} ---\n{content[:3000]}\n"
        
        #limit content
        combined_content = combined_content[:15000]
        
        # gemini models 
        models = ["gemini-2.5-flash", "gemini-2.5-flash-lite"]
        model_name = models[min(retry_count, len(models)-1)]
        
        # prompt
        prompt = f"""Extract NGO information and return ONLY valid JSON. No explanations, no markdown.

Website: {url}

Content:
{combined_content}

Return this EXACT JSON structure (copy these field names exactly):
{{
  "NGO Name": "official organization name",
  "Address": "complete physical address with city, state, pincode",
  "Services Offered": "service1; service2; service3",
  "Contact Person Details": "name or email of contact person",
  "Contact Number": "phone with country code like +91 XXXXX XXXXX"
}}

Rules:
- NGO Name: Extract official name only, no extra words
- Address: Must have street, area, city, state, PIN code if available
- Services Offered: List 3-5 main services separated by semicolons
- Contact Person Details: Name of founder/director OR email address
- Contact Number: Format as +91 XXXXX XXXXX for Indian numbers
- Use "Not found" ONLY if truly not present in any section

Return ONLY the JSON object, nothing else."""
        
        model = genai.GenerativeModel(model_name)
        
        response = model.generate_content(
            prompt,
            generation_config=genai.GenerationConfig(
                max_output_tokens=1500,
                temperature=0.05, 
                top_p=0.85,
                top_k=40
            ),
            safety_settings={
                'HATE': 'BLOCK_NONE',
                'HARASSMENT': 'BLOCK_NONE',
                'SEXUAL': 'BLOCK_NONE',
                'DANGEROUS': 'BLOCK_NONE'
            }
        )
        
        # Extract response text with multiple fall backs
        response_text = None
        
        if hasattr(response, 'text'):
            response_text = response.text
        elif hasattr(response, 'candidates') and response.candidates:
            candidate = response.candidates[0]
            
            if candidate.finish_reason not in [1, "STOP"]:
                # If blocked or incomplete, retry with different model
                if retry_count < 2:
                    logger.warning(f"Retry {retry_count + 1}: finish_reason={candidate.finish_reason}")
                    time.sleep(1)
                    return extract_required_fields_with_gemini(all_content, url, retry_count + 1)
                else:
                    return {
                        "NGO Name": f"  Generation blocked (reason: {candidate.finish_reason})",
                        "Address": "Not found",
                        "Services Offered": "Not found",
                        "Contact Person Details": "Not found",
                        "Contact Number": "Not found"
                    }
            
            if hasattr(candidate, 'content') and hasattr(candidate.content, 'parts') and candidate.content.parts:
                response_text = candidate.content.parts[0].text
        
        if not response_text:
            if retry_count < 2:
                logger.warning(f"Retry {retry_count + 1}: No response text")
                time.sleep(1)
                return extract_required_fields_with_gemini(all_content, url, retry_count + 1)
            else:
                return {
                    "NGO Name": "  No response from AI",
                    "Address": "Not found",
                    "Services Offered": "Not found",
                    "Contact Person Details": "Not found",
                    "Contact Number": "Not found"
                }
        
        extracted_data = clean_json_response(response_text)
        
        if not extracted_data:
            if retry_count < 2:
                logger.warning(f"Retry {retry_count + 1}: JSON parsing failed")
                time.sleep(1)
                return extract_required_fields_with_gemini(all_content, url, retry_count + 1)
            else:
                logger.error(f"Failed to parse JSON after {retry_count + 1} attempts. Response: {response_text[:200]}")
                return {
                    "NGO Name": "  AI returned invalid JSON after 3 attempts",
                    "Address": "Not found",
                    "Services Offered": "Not found",
                    "Contact Person Details": "Not found",
                    "Contact Number": "Not found"
                }
            
        field_mapping = {
            "ngo name": "NGO Name",
            "name": "NGO Name",
            "organization name": "NGO Name",
            "address": "Address",
            "location": "Address",
            "services offered": "Services Offered",
            "services": "Services Offered",
            "contact person details": "Contact Person Details",
            "contact person": "Contact Person Details",
            "contact": "Contact Person Details",
            "contact number": "Contact Number",
            "phone": "Contact Number",
            "phone number": "Contact Number",
            "telephone": "Contact Number"
        }
        
        normalized_data = {}
        for key, value in extracted_data.items():
            normalized_key = key.strip()
            # Find matching standard field name
            for alt_key, standard_key in field_mapping.items():
                if normalized_key.lower() == alt_key:
                    normalized_key = standard_key
                    break
            normalized_data[normalized_key] = value
        
        #Ensures all required fields are present
        required_fields = ["NGO Name", "Address", "Services Offered", "Contact Person Details", "Contact Number"]
        for field in required_fields:
            if field not in normalized_data:
                normalized_data[field] = "Not found"
            elif not normalized_data[field] or str(normalized_data[field]).strip().lower() in ["null", "none", "", "n/a", "na", "nil", "not available"]:
                normalized_data[field] = "Not found"
            else:
                normalized_data[field] = str(normalized_data[field]).strip()
        
        # Check if too many fields are "Not found"
        not_found_count = sum(1 for v in normalized_data.values() if v == "Not found")
        
        #REtry if more than 3 fields anre not found
        if not_found_count > 3 and retry_count < 1:
            logger.warning(f"Retry {retry_count + 1}: Too many fields not found ({not_found_count}/5)")
            time.sleep(1)
            return extract_required_fields_with_gemini(all_content, url, retry_count + 1)
        
        return normalized_data
        
    except json.JSONDecodeError as e:
        if retry_count < 2:
            logger.warning(f"Retry {retry_count + 1}: JSON decode error")
            time.sleep(1)
            return extract_required_fields_with_gemini(all_content, url, retry_count + 1)
        else:
            logger.error(f"JSON parsing failed after retries: {str(e)}")
            return {
                "NGO Name": "  AI returned invalid JSON",
                "Address": "Not found",
                "Services Offered": "Not found",
                "Contact Person Details": "Not found",
                "Contact Number": "Not found"
            }
    except ImportError:
        return {
            "NGO Name": "  google-generativeai not installed",
            "Address": "Run: pip install google-generativeai",
            "Services Offered": "Not found",
            "Contact Person Details": "Not found",
            "Contact Number": "Not found"
        }
    except Exception as e:
        if retry_count < 2:
            logger.warning(f"Retry {retry_count + 1}: Unexpected error: {str(e)}")
            time.sleep(1)
            return extract_required_fields_with_gemini(all_content, url, retry_count + 1)
        else:
            logger.error(f"Extraction failed after retries: {str(e)}")
            return {
                "NGO Name": f"  AI Error: {str(e)[:100]}",
                "Address": "Not found",
                "Services Offered": "Not found",
                "Contact Person Details": "Not found",
                "Contact Number": "Not found"
            }

def scrape_and_extract_ngo_data(url):
    """Main function: Comprehensive scraping + AI extraction with retries"""
    all_content, final_url = scrape_comprehensive_content(url)
    
    if all_content and all_content[0][0] == "Error":
        return {
            "NGO Name": all_content[0][1],
            "Address": "Not found",
            "Services Offered": "Not found",
            "Contact Person Details": "Not found",
            "Contact Number": "Not found"
        }
    
    extracted_data = extract_required_fields_with_gemini(all_content, final_url)
    return extracted_data

# Streamlit UI
st.markdown('<h1 class="main-header">🌐 NGO Web Scraper <span class="ai-badge">✨ Gemini AI</span></h1>', unsafe_allow_html=True)



#Check for gemini api key 
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")
if GEMINI_API_KEY:
    st.success("✅ System is ready!")
else:
    st.warning("⚠️ **Gemini API key not found!** Create a `.env` file with `GEMINI_API_KEY=your_api_key_here`")
    st.info("📝 Get your free API key from: https://makersuite.google.com/app/apikey")

# Input section
col1, col2 = st.columns([3, 1])
with col1:
    url_input = st.text_input("Enter NGO website URL", placeholder="https://example-ngo.org")
with col2:
    st.write("")
    st.write("")
    scrape_button = st.button("🚀 Smart Scrape")


if scrape_button and url_input:
    with st.spinner("🔄 Scanning multiple pages... 🤖 AI extracting with auto-retry..."):
        if not url_input.startswith(('http://', 'https://')):
            url_input = 'https://' + url_input
        
        scraped_data = scrape_and_extract_ngo_data(url_input)
        st.session_state.scraped_data = [scraped_data]
        st.session_state.current_url = url_input
        
        #Count fields that are not found
        success_count = sum(1 for v in scraped_data.values() if v != "Not found")

        is_error = scraped_data["NGO Name"].startswith("  ")

        if is_error:
            st.error("  Extraction failed - check error details below")
        elif success_count == 5:
            st.success("🎉 Perfect! Extracted all 5/5 fields successfully!")
        elif success_count >= 4:
            st.success(f"✅ Excellent! Extracted {success_count}/5 fields successfully!")
        elif success_count >= 3:
            st.info(f"ℹ️ Good! Extracted {success_count}/5 fields. Some data may not be available on the website.")
        else:
            st.warning(f"⚠️ Partial extraction: {success_count}/5 fields found")
        
        st.subheader("📋 Extracted Required Fields")
        
        for key, value in scraped_data.items():
            
            is_error = value.startswith("  ")
            is_not_found = value == "Not found"
            is_success = not is_error and not is_not_found
            
            if is_error:
                display_value = f'<span style="color: #d32f2f;">{value}</span>'
                badge = ""
            elif is_not_found:
                display_value = '<span class="not-found">Not found on website</span>'
                badge = ""
            else:
                display_value = value
                badge = ' <span class="success-badge">✓</span>'
            
            st.markdown(f"""
            <div class="data-card">
                <div class="field-label">{key}{badge}</div>
                <div class="field-value">{display_value}</div>
            </div>
            """, unsafe_allow_html=True)


if st.session_state.scraped_data:
    st.markdown("---")
    st.subheader("💾 Download Results")
    
    excel_file = create_excel_file(st.session_state.scraped_data)
    
    if excel_file:
        col1, col2 = st.columns([1, 3])
        with col1:
            st.download_button(
                label="📥 Download Excel",
                data=excel_file,
                file_name="ngo_data_complete.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )



